from openpyxl import load_workbook


wb_pr = load_workbook(r'exel\report_2021.xlsx')
sheet = wb_pr.active
quantity_row = sheet.max_row
for i in range(2, quantity_row + 1):
    name = str(sheet['A' + str(i)].value)
    value = str(sheet['B' + str(i)].value)
    print(name, ' ', value)
