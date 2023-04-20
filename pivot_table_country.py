import decimal
import openpyxl
import win32com.client
from settings import path_exel, path_exel2, list_year, rate, year_now
import pandas as pd
import os
from openpyxl import load_workbook


ctx = decimal.Context()


def pivot1_1(year):
    wb = load_workbook(path_exel)
    sheet = wb[year]
    sheet['K1'] = 'Страны'
    sheet['L1'] = '№_ФО'
    sheet['M1'] = '№_Региона'
    sheet['N1'] = 'ФО'
    sheet['O1'] = 'Регионы'
    sheet['P1'] = 'Цена_д'
    sheet['Q1'] = 'Цена_р'

    wb.save(path_exel)

    # количество строк
    quantity_row = sheet.max_row

    # заменяем запятую на точку и сохраняем на новый лист
    data_frame = pd.ExcelFile(path_exel)
    data_frame_last = data_frame.parse(year, decimal=',')
    with pd.ExcelWriter(path_exel2) as writer:
        data_frame_last.to_excel(writer, 'last')

    # удаляем столбец
    wb_2 = openpyxl.load_workbook(path_exel2)
    sheet_2 = wb_2.active
    sheet_2.delete_cols(1, 1)
    wb_2.save(path_exel2)

    # вставляем лист со "странами", ФО и регионами
    path_1 = r'exel\1_1.xlsx'
    path_fo = r'exel\1_2_fo.xlsx'
    path_reg = r'exel\1_3_reg.xlsx'
    wb_2.create_sheet('Prob')
    wb_2.create_sheet('FO')
    wb_2.create_sheet('Reg')

    # вставляем лист со "странами"
    insert_sheet(path_1, wb_2, path_exel2, 'Prob')
    insert_sheet(path_fo, wb_2, path_exel2, 'FO')
    insert_sheet(path_reg, wb_2, path_exel2, 'Reg')

    # вставляем формулу со "странами"
    wb_form = load_workbook(path_exel2)
    sheet_form = wb_form['last']
    for k in range(2, quantity_row + 1):
        formula = '=VLOOKUP(C' + str(k) + ',Prob!A:B,2,0)'
        formula2 = '=F' + str(k) + '/G' + str(k)
        formula3 = r'=P' + str(k) + '*' + str(rate[year_now])

        sheet_form.cell(row=k, column=11, value=formula)
        sheet_form.cell(row=k, column=16, value=formula2)
        sheet_form.cell(row=k, column=17, value=formula3)
    wb_form.save(path_exel2)

    # вставляем формулу с "номерами федеральных округов"
    wb_form = load_workbook(path_exel2)
    sheet_form = wb_form['last']
    for k in range(2, quantity_row + 1):
        formula = '=LEFT(J' + str(k) + ',2)'
        sheet_form.cell(row=k, column=12, value=formula)
    wb_form.save(path_exel2)

    # вставляем формулу с "номерами регионов"
    wb_form = load_workbook(path_exel2)
    sheet_form = wb_form['last']
    for k in range(2, quantity_row + 1):
        formula = '=LEFT(I' + str(k) + ',5)'
        sheet_form.cell(row=k, column=13, value=formula)
    wb_form.save(path_exel2)

    # вставляем формулу с "именами федеральных округов"
    wb_form = load_workbook(path_exel2)
    sheet_form = wb_form['last']
    for k in range(2, quantity_row + 1):
        formula = '=VLOOKUP(L' + str(k) + ',FO!A:B,2,0)'
        sheet_form.cell(row=k, column=14, value=formula)
    wb_form.save(path_exel2)

    # вставляем формулу с "именами регионов"
    wb_form = load_workbook(path_exel2)
    sheet_form = wb_form['last']
    for k in range(2, quantity_row + 1):
        formula = '=VLOOKUP(M' + str(k) + ',Reg!A:B,2,0)'
        sheet_form.cell(row=k, column=15, value=formula)
    wb_form.save(path_exel2)
    pass


def pivot1_2():
    # вычисляем сводную таблицу по странам
    excel_path = r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"
    excel_file_path = path_exel2

    excel_file = os.path.join(excel_path, excel_file_path)
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False
    excel.Workbooks.Open(excel_file)
    excel.ActiveWorkbook.SaveAs(excel_file, FileFormat=51, ConflictResolution=2)
    excel.DisplayAlerts = True
    excel.ActiveWorkbook.Close()

    data_frame_country = pd.read_excel(path_exel2)

    report_table = data_frame_country.pivot_table(index='Страны', values='NETTO', aggfunc='sum').round(0)
    report_table.to_excel(r'exel\report_2021.xlsx', sheet_name='Report')

    df = pd.read_excel(r'exel\report_2021.xlsx')
    bit_depth = 0
    col_netto_list = df['NETTO'].tolist()
    exist_count = col_netto_list.count(0)

    while exist_count != 0:
        if bit_depth > 6:
            break
        else:
            bit_depth += 1
            report_table = data_frame_country.pivot_table(index='Страны', values='NETTO', aggfunc='sum').round(bit_depth)
            report_table.to_excel(r'exel\report_2021.xlsx',
                                  sheet_name='Report')
            df = pd.read_excel(r'exel\report_2021.xlsx')
            col_netto_list = df['NETTO'].tolist()
            exist_count = col_netto_list.count(0)

    # сортировка таблицы по нетто
    df_netto = pd.read_excel(r'exel\report_2021.xlsx')
    final_result = df_netto.sort_values(by='NETTO', ascending=False)
    final_result.to_excel(r'exel\report_2021.xlsx',
                          sheet_name='Report')

    # удаляем столбец
    wb_netto = openpyxl.load_workbook(r'exel\report_2021.xlsx')
    sheet_3 = wb_netto.active
    sheet_3.delete_cols(1, 1)
    wb_netto.save(r'exel\report_2021.xlsx')

    # добавляем "итого" по странам
    df_country_itog = pd.read_excel(r'exel\report_2021.xlsx')
    itog = round(df_country_itog["NETTO"].sum())
    wb_report = load_workbook(r'exel\report_2021.xlsx')
    sheet = wb_report.active
    quantity_row = sheet.max_row
    sheet['A' + str(quantity_row + 1)] = 'Итого'
    sheet['B' + str(quantity_row + 1)] = itog
    wb_report.save(r'exel\report_2021.xlsx')

    # добавляем название столбца "Доля"
    wb_proportion = load_workbook(r'exel\report_2021.xlsx')
    sheet_proportion = wb_proportion.active
    sheet_proportion['C1'] = 'Доля'

    # расчитываем долю по каждой стране
    counter_country = 1
    counter_netto = 0
    quantity_row = sheet_proportion.max_row
    for i_netto in range(2, quantity_row + 1):
        netto_country = sheet_proportion['B' + str(i_netto)].value
        netto_proportion = round(netto_country / itog * 100, counter_country)

        netto_value = round(sheet_proportion['B' + str(i_netto)].value)

        while netto_value == 0:
            if counter_netto > 6:
                break
            else:
                counter_netto += 1
                netto_value = round(sheet_proportion['B' + str(i_netto)].value, counter_netto)
        sheet_proportion['B' + str(i_netto)].value = netto_value

        while netto_proportion == 0:
            if counter_country > 6:
                break
            else:
                counter_country += 1
                netto_proportion = round(netto_country / itog * 100, counter_country)
        sheet_proportion['C' + str(i_netto)].value = netto_proportion

    quantity_row = sheet_proportion.max_row
    sheet_proportion['C' + str(quantity_row)] = round(100, 1)
    wb_proportion.save(r'exel\report_2021.xlsx')

    # суммируем доли  и объем в странах, доли которых менее 0,01%
    wb_proportion_part = load_workbook(r'exel\report_2021.xlsx')
    sheet_proportion_part = wb_proportion_part.active
    quantity_row = sheet_proportion_part.max_row
    proportion_row = []
    proportion_part_value = 0
    proportion_netto_value = 0
    for i_part in range(2, quantity_row):
        proportion_part_country = sheet_proportion_part['C' + str(i_part)].value
        netto_part_country = sheet_proportion_part['B' + str(i_part)].value

        if proportion_part_country < 0.01:
            proportion_part_value += float(proportion_part_country)
            proportion_row.append(i_part)
            proportion_netto_value += netto_part_country

    if proportion_part_value != 0:
        # удаляем строки, доли которых менее 0,01%
        number = 0
        for number_row in proportion_row:
            sheet_proportion_part.delete_rows(number_row - number)
            number += 1
        quantity_row = sheet_proportion_part.max_row

        # вставляем пустую строку и пишем значение "другие"
        sheet_proportion_part.insert_rows(sheet_proportion_part.max_row, 1)
        sheet_proportion_part['A' + str(quantity_row)] = 'другие'

        if proportion_netto_value >= 1:
            proportion_netto_value = round(proportion_netto_value, 0)
            sheet_proportion_part['B' + str(quantity_row)] = proportion_netto_value
        elif proportion_netto_value == 0:
            pass
        else:
            proportion_netto_value_ = round(proportion_netto_value, 0)
            number4 = 0
            while proportion_netto_value_ == 0:
                number4 += 1
                proportion_netto_value_ = round(proportion_netto_value, number4)

            proportion_netto_value = proportion_netto_value_
            sheet_proportion_part['B' + str(quantity_row)] = proportion_netto_value

        precision = 6
        proportion_part_value = f'{proportion_part_value:.{precision}f}'

        list_number = ['1', '2', '3', '4', '5', '6', '7', '8', '9']
        index_list = []
        for number in list_number:
            if number in proportion_part_value:
                index = proportion_part_value.find(number)
                index_list.append(index)
        min_index_list = min(index_list)
        proportion_part_value = proportion_part_value[0:min_index_list + 1]
        sheet_proportion_part['C' + str(quantity_row)] = proportion_part_value

    wb_proportion_part.save(r'exel\report_2021.xlsx')
    pass


# вставляем сводную таблицу из экселя в ворд и расчитываем долю четырех крупнейших стран
def pivot1_3():
    table_contents_country = []
    table_contents_country_int = []
    workbook = load_workbook(r'exel\report_2021.xlsx')
    sheet_1 = workbook['Report']
    for i in range(2, sheet_1.max_row + 1):
        table_contents_country_int.append({
            'Страны': sheet_1.cell(i, 1).value,
            'NETTO': sheet_1.cell(i, 2).value,
            'Доля': sheet_1.cell(i, 3).value
        })
        d = decimal.Decimal(sheet_1.cell(i, 2).value)
        d = moneyfmt(d, sep=' ')
        table_contents_country.append({
            'Страны': sheet_1.cell(i, 1).value,
            'NETTO': d,
            'Доля': sheet_1.cell(i, 3).value
        })
    # сумма долей 4-х крупнейших стран
    country_part = 0
    len_table_contents_country = len(table_contents_country)
    if len(table_contents_country) == 2:
        country_part = float(table_contents_country[0]['Доля'])
        # table_contents_country.append({'Страны': 'нет данных', 'NETTO': 0, 'Доля': 0})
        # table_contents_country.append({'Страны': 'нет данных', 'NETTO': 0, 'Доля': 0})
        # table_contents_country.append({'Страны': 'нет данных', 'NETTO': 0, 'Доля': 0})

    elif len(table_contents_country) == 3:
        country_part = float(table_contents_country[0]['Доля']) + float(table_contents_country[1]['Доля'])
        # table_contents_country.append({'Страны': 'нет данных', 'NETTO': 0, 'Доля': 0})
        # table_contents_country.append({'Страны': 'нет данных', 'NETTO': 0, 'Доля': 0})

    elif len(table_contents_country) == 4:
        country_part = float(table_contents_country[0]['Доля']) + float(table_contents_country[1]['Доля']) + float(
            table_contents_country[2]['Доля'])
        # table_contents_country.append({'Страны': 'нет данных', 'NETTO': 0, 'Доля': 0})

    else:
        country_part = float(table_contents_country[0]['Доля']) + float(table_contents_country[1]['Доля']) + float(
            table_contents_country[2]['Доля']) + float(table_contents_country[3]['Доля'])

    return table_contents_country, country_part, table_contents_country_int, len_table_contents_country


# вставить пробелы между разрадами в длинных числах (123456 = 123 456)
def moneyfmt(value, places=0, curr='', sep=',', dp='',
             pos='', neg='-', trailneg=''):
    q = decimal.Decimal(10) ** -places      # 2 places --> '0.01'
    sign, digits, exp = value.quantize(q).as_tuple()
    result = []
    digits = list(map(str, digits))
    build = result.append
    next2 = digits.pop
    if sign:
        build(trailneg)
    for i in range(places):
        build(next2() if digits else '0')
    build(dp)
    if not digits:
        build('0')
    i = 0
    while digits:
        build(next2())
        i += 1
        if i == 3 and digits:
            i = 0
            build(sep)
    build(curr)
    build(neg if sign else pos)
    return ''.join(reversed(result))


# сводная таблица по странам и стоимости
def pivot2_1():
    excel_path = r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"
    excel_file_path = path_exel2

    excel_file = os.path.join(excel_path, excel_file_path)
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False
    excel.Workbooks.Open(excel_file)
    excel.ActiveWorkbook.SaveAs(excel_file, FileFormat=51, ConflictResolution=2)
    excel.DisplayAlerts = True
    excel.ActiveWorkbook.Close()

    data_frame_country = pd.read_excel(path_exel2)

    report_table = data_frame_country.pivot_table(index='Страны', values='STOIM', aggfunc='sum').round(0)
    report_table.to_excel(r'exel\report_2021.xlsx',
                          sheet_name='Report')

    df = pd.read_excel(r'exel\report_2021.xlsx')
    bit_depth = 0
    col_netto_list = df['STOIM'].tolist()
    exist_count = col_netto_list.count(0)

    while exist_count != 0:
        if bit_depth > 6:
            break
        else:
            bit_depth += 1
            report_table = data_frame_country.pivot_table(index='Страны', values='STOIM', aggfunc='sum').round(bit_depth)
            report_table.to_excel(r'exel\report_2021.xlsx',
                                  sheet_name='Report')
            df = pd.read_excel(r'exel\report_2021.xlsx')
            col_netto_list = df['STOIM'].tolist()
            exist_count = col_netto_list.count(0)

    # сортировка таблицы по стоимости
    df_netto = pd.read_excel(r'exel\report_2021.xlsx')
    final_result = df_netto.sort_values(by='STOIM', ascending=False)
    final_result.to_excel(r'exel\report_2021.xlsx',
                          sheet_name='Report')

    # удаляем столбец
    wb_netto = openpyxl.load_workbook(r'exel\report_2021.xlsx')
    sheet_3 = wb_netto.active
    sheet_3.delete_cols(1, 1)
    wb_netto.save(r'exel\report_2021.xlsx')

    # добавляем "итого" по странам
    df_country_itog = pd.read_excel(r'exel\report_2021.xlsx')
    itog = round(df_country_itog["STOIM"].sum())
    wb_report = load_workbook(r'exel\report_2021.xlsx')
    sheet = wb_report.active
    quantity_row = sheet.max_row
    sheet['A' + str(quantity_row + 1)] = 'Итого'
    sheet['B' + str(quantity_row + 1)] = itog
    wb_report.save(r'exel\report_2021.xlsx')

    # добавляем название столбца "Доля"
    wb_proportion = load_workbook(r'exel\report_2021.xlsx')
    sheet_proportion = wb_proportion.active
    sheet_proportion['C1'] = 'Доля'

    # расчитываем долю по каждой стране
    counter_country = 1
    counter_netto = 0
    quantity_row = sheet_proportion.max_row
    for i_netto in range(2, quantity_row + 1):
        netto_country = sheet_proportion['B' + str(i_netto)].value
        netto_proportion = round(netto_country / itog * 100, counter_country)

        netto_value = round(sheet_proportion['B' + str(i_netto)].value)

        while netto_value == 0:
            if counter_netto > 6:
                break
            else:
                counter_netto += 1
                netto_value = round(sheet_proportion['B' + str(i_netto)].value, counter_netto)
        sheet_proportion['B' + str(i_netto)].value = netto_value

        while netto_proportion == 0:
            if counter_country > 6:
                break
            else:
                counter_country += 1
                netto_proportion = round(netto_country / itog * 100, counter_country)
        sheet_proportion['C' + str(i_netto)].value = netto_proportion

    quantity_row = sheet_proportion.max_row
    sheet_proportion['C' + str(quantity_row)] = round(100, 1)
    wb_proportion.save(r'exel\report_2021.xlsx')

    # суммируем доли  и объем в странах, доли которых менее 0,01%
    wb_proportion_part = load_workbook(r'exel\report_2021.xlsx')
    sheet_proportion_part = wb_proportion_part.active
    quantity_row = sheet_proportion_part.max_row
    proportion_row = []
    proportion_part_value = 0
    proportion_netto_value = 0
    for i_part in range(2, quantity_row):
        proportion_part_country = sheet_proportion_part['C' + str(i_part)].value
        netto_part_country = sheet_proportion_part['B' + str(i_part)].value

        if proportion_part_country < 0.01:
            proportion_part_value += float(proportion_part_country)
            proportion_row.append(i_part)
            proportion_netto_value += netto_part_country

    if proportion_netto_value != 0:
        # удаляем строки, доли которых менее 0,01%
        number = 0
        for number_row in proportion_row:
            sheet_proportion_part.delete_rows(number_row - number)
            number += 1
        quantity_row = sheet_proportion_part.max_row

        # вставляем пустую строку и пишем значение "другие"
        sheet_proportion_part.insert_rows(sheet_proportion_part.max_row, 1)
        sheet_proportion_part['A' + str(quantity_row)] = 'другие'

        if proportion_netto_value >= 1:
            proportion_netto_value = round(proportion_netto_value, 0)
            sheet_proportion_part['B' + str(quantity_row)] = proportion_netto_value
        else:
            proportion_netto_value = round(proportion_netto_value, 0)
            number4 = 0
            while proportion_netto_value == 0:
                if number4 > 6:
                    break
                else:
                    number4 += 1
                    proportion_netto_value = round(proportion_netto_value, number4)
            sheet_proportion_part['B' + str(quantity_row)] = proportion_netto_value

        precision = 6
        proportion_part_value = f'{proportion_part_value:.{precision}f}'

        list_number = ['1', '2', '3', '4', '5', '6', '7', '8', '9']
        index_list = []
        for number in list_number:
            if number in proportion_part_value:
                index = proportion_part_value.find(number)
                index_list.append(index)
        min_index_list = min(index_list)
        proportion_part_value = proportion_part_value[0:min_index_list + 1]
        sheet_proportion_part['C' + str(quantity_row)] = proportion_part_value

    wb_proportion_part.save(r'exel\report_2021.xlsx')
    pass


# вставляем сводную таблицу из экселя в ворд и расчитываем долю четырех крупнейших стран в стоимостном выражении
def pivot2_2():
    table_contents_country = []
    workbook = load_workbook(r'exel\report_2021.xlsx')
    sheet_1 = workbook['Report']
    for i in range(2, sheet_1.max_row + 1):
        d = decimal.Decimal(sheet_1.cell(i, 2).value)
        d = moneyfmt(d, sep=' ')
        table_contents_country.append({
            'Страны': sheet_1.cell(i, 1).value,
            'STOIM': d,
            'Доля': sheet_1.cell(i, 3).value
        })
    # сумма долей 4-х крупнейших стран
    if len(table_contents_country) == 2:
        country_part = float(table_contents_country[0]['Доля'])
    elif len(table_contents_country) == 3:
        country_part = float(table_contents_country[0]['Доля']) + float(table_contents_country[1]['Доля'])
    elif len(table_contents_country) == 4:
        country_part = float(table_contents_country[0]['Доля']) + float(table_contents_country[1]['Доля']) + float(
            table_contents_country[2]['Доля'])
    else:
        country_part = float(table_contents_country[0]['Доля']) + float(table_contents_country[1]['Доля']) + float(
        table_contents_country[2]['Доля'] + float(table_contents_country[3]['Доля']))

    return table_contents_country, country_part


def pivot_table_country_year():
    pivot1_1('2013')
    try:
        pivot1_2()
        df = pd.read_excel(r'exel\report_2021.xlsx')
        df.pop('Доля')
        df.drop(labels=[len(df) - 1], axis=0, inplace=True)
        df.loc[:, "Год"] = "2013"
        df.to_excel(r'exel\report_2021_all.xlsx', index=False)
    except Exception:
        print('Пропуск 2013 г. в pivot_table_country_year')
        df5 = pd.DataFrame()
        df5.to_excel(r'exel\report_2021_all.xlsx', index=False)

    for year in range(1, len(list_year)):
        wb_report_now = load_workbook(r'exel\report_2021_all.xlsx')
        sheet_now = wb_report_now.active
        row_end_now = sheet_now.max_row
        if row_end_now > 1:
            pivot1_1(list_year[year])
            try:
                pivot1_2()
                df = pd.read_excel(r'exel\report_2021.xlsx')
                df.pop('Доля')
                df.drop(labels=[len(df) - 1], axis=0, inplace=True)
                df.loc[:, "Год"] = list_year[year]
                wb_report = load_workbook(r'exel\report_2021_all.xlsx')
                sheet = wb_report.active
                row_end = sheet.max_row
                with pd.ExcelWriter(r'exel\report_2021_all.xlsx', mode='a', if_sheet_exists='overlay') as writer:
                    df.to_excel(writer, startrow=row_end, index=False, header=False)

            except Exception:
                print(f'Пропуск {list_year[year]} г. в pivot_table_country_year')
        else:
            pivot1_1(list_year[year])
            try:
                pivot1_2()
                df = pd.read_excel(r'exel\report_2021.xlsx')
                df.pop('Доля')
                df.drop(labels=[len(df) - 1], axis=0, inplace=True)
                df.loc[:, "Год"] = list_year[year]
                df.to_excel(r'exel\report_2021_all.xlsx', index=False)

            except Exception:
                print(f'Пропуск {list_year[year]} г. в pivot_table_country_year')

    data_frame_country = pd.read_excel(r'exel\report_2021_all.xlsx')
    report_table = data_frame_country.pivot_table(index='Страны', values='NETTO', aggfunc='sum', columns='Год',
                                                  margins=True).round(0)
    report_table.to_excel(r'exel\report_2021.xlsx', sheet_name='Report')

    table_contents_country = []
    workbook = load_workbook(r'exel\report_2021.xlsx')
    sheet_1 = workbook['Report']
    for i in range(2, sheet_1.max_row + 1):
        table_year_dict = dict()
        table_year_dict['Страны'] = sheet_1.cell(i, 1).value
        for j in range(2, sheet_1.max_column + 1):
            d = sheet_1.cell(i, j).value
            if d is not None:
                d = decimal.Decimal(sheet_1.cell(i, j).value)
                d = moneyfmt(d, sep=' ')
                table_year_dict[sheet_1.cell(1, j).value] = d
        table_contents_country.append(table_year_dict)

    return table_contents_country


def insert_sheet(path, wb, path_exel_2, sheet_name):
    wb_1 = openpyxl.load_workbook(path)
    sheet_1 = wb_1.active
    quantity_row_1 = sheet_1.max_row
    quantity_column_1 = sheet_1.max_column
    for i in range(1, quantity_row_1 + 1):
        for j in range(1, quantity_column_1 + 1):
            a = sheet_1.cell(row=i, column=j)
            b = a.value
            sheet2 = wb[sheet_name]
            sheet2.cell(row=i, column=j, value=b)
    wb.save(path_exel_2)


# сводная таблица по федеральным округам, регионам в тоннаже
def pivot_fo_1(fo_reg, netto_stoim, path_save):
    excel_path = r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"
    excel_file_path = path_exel2

    excel_file = os.path.join(excel_path, excel_file_path)
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False
    excel.Workbooks.Open(excel_file)
    excel.ActiveWorkbook.SaveAs(excel_file, FileFormat=51, ConflictResolution=2)
    excel.DisplayAlerts = True
    excel.ActiveWorkbook.Close()

    data_frame_country = pd.read_excel(path_exel2)

    report_table = data_frame_country.pivot_table(index=fo_reg, values=netto_stoim, aggfunc='sum').round(0)
    report_table.to_excel(path_save, sheet_name='Report')

    df = pd.read_excel(path_save)
    bit_depth = 0
    col_netto_list = df[netto_stoim].tolist()
    exist_count = col_netto_list.count(0)

    while exist_count != 0:
        if bit_depth > 6:
            break
        else:
            bit_depth += 1
            report_table = data_frame_country.pivot_table(index=fo_reg, values=netto_stoim, aggfunc='sum').\
                round(bit_depth)
            report_table.to_excel(path_save, sheet_name='Report')
            df = pd.read_excel(path_save)
            col_netto_list = df[netto_stoim].tolist()
            exist_count = col_netto_list.count(0)

    # сортировка таблицы по весу
    df_netto = pd.read_excel(path_save)
    final_result = df_netto.sort_values(by=netto_stoim, ascending=False)
    final_result.to_excel(path_save, sheet_name='Report')

    # удаляем столбец
    wb_netto = openpyxl.load_workbook(path_save)
    sheet_3 = wb_netto.active
    sheet_3.delete_cols(1, 1)
    wb_netto.save(path_save)

    # добавляем "итого" по федеральным округам
    df_country_itog_last = pd.read_excel(path_exel2)
    itog = round(df_country_itog_last[netto_stoim].sum())
    wb_report = load_workbook(path_save)
    sheet = wb_report.active
    quantity_row = sheet.max_row
    sheet['A' + str(quantity_row + 1)] = 'Итого'
    sheet['B' + str(quantity_row + 1)] = itog
    wb_report.save(path_save)

    # добавляем название столбца "Доля"
    wb_proportion = load_workbook(path_save)
    sheet_proportion = wb_proportion.active
    sheet_proportion['C1'] = 'Доля'

    # расчитываем долю по каждому ФО
    counter_country = 1
    counter_netto = 0
    quantity_row = sheet_proportion.max_row
    for i_netto in range(2, quantity_row + 1):
        netto_country = sheet_proportion['B' + str(i_netto)].value
        netto_proportion = round(netto_country / itog * 100, counter_country)

        netto_value = round(sheet_proportion['B' + str(i_netto)].value)

        while netto_value == 0:
            if counter_netto > 6:
                break
            else:
                counter_netto += 1
                netto_value = round(sheet_proportion['B' + str(i_netto)].value, counter_netto)
        sheet_proportion['B' + str(i_netto)].value = netto_value

        while netto_proportion == 0:
            if counter_country > 6:
                break
            else:
                counter_country += 1
                netto_proportion = round(netto_country / itog * 100, counter_country)
        sheet_proportion['C' + str(i_netto)].value = netto_proportion

    quantity_row = sheet_proportion.max_row
    sheet_proportion['C' + str(quantity_row)] = round(100, 1)
    wb_proportion.save(path_save)

    # суммируем доли  и объем в странах, доли которых менее 0,01%
    wb_proportion_part = load_workbook(path_save)
    sheet_proportion_part = wb_proportion_part.active
    quantity_row = sheet_proportion_part.max_row
    proportion_row = []
    proportion_part_value = 0
    proportion_netto_value = 0
    for i_part in range(2, quantity_row):
        proportion_part_country = sheet_proportion_part['C' + str(i_part)].value
        netto_part_country = sheet_proportion_part['B' + str(i_part)].value

        if proportion_part_country < 0.01:
            proportion_part_value += float(proportion_part_country)
            proportion_row.append(i_part)
            proportion_netto_value += netto_part_country

    # удаляем строки, доли которых менее 0,01%
    number = 0
    for number_row in proportion_row:
        sheet_proportion_part.delete_rows(number_row - number)
        number += 1
    quantity_row = sheet_proportion_part.max_row

    # вставляем пустую строку и пишем значение "другие"
    if len(proportion_row) != 0:
        sheet_proportion_part.insert_rows(sheet_proportion_part.max_row, 1)
        sheet_proportion_part['A' + str(quantity_row)] = 'другие'

        if proportion_netto_value >= 1:
            proportion_netto_value = round(proportion_netto_value, 0)
            sheet_proportion_part['B' + str(quantity_row)] = proportion_netto_value
        else:
            proportion_netto_value = round(proportion_netto_value, 0)
            number4 = 0
            while proportion_netto_value == 0:
                if number4 > 6:
                    break
                else:
                    number4 += 1
                    proportion_netto_value = round(proportion_netto_value, number4)
            sheet_proportion_part['B' + str(quantity_row)] = proportion_netto_value

        precision = 6
        proportion_part_value = f'{proportion_part_value:.{precision}f}'

        list_number = ['1', '2', '3', '4', '5', '6', '7', '8', '9']
        index_list = []
        for number in list_number:
            if number in proportion_part_value:
                index = proportion_part_value.find(number)
                index_list.append(index)
        min_index_list = min(index_list)
        proportion_part_value = proportion_part_value[0:min_index_list + 1]
        sheet_proportion_part['C' + str(quantity_row)] = proportion_part_value
    wb_proportion_part.save(path_save)
    pass


# вставляем сводную таблицу по федеральным округам и регионам из экселя в ворд и расчитываем долю четырех крупнейших ФО
def pivotfo(path_save1, fo_reg, netto_stoim):
    table_contents_country = []
    table_contents_country_int = []
    workbook = load_workbook(path_save1)
    sheet_1 = workbook['Report']
    for i in range(2, sheet_1.max_row + 1):
        table_contents_country_int.append({
            fo_reg: sheet_1.cell(i, 1).value,
            netto_stoim: sheet_1.cell(i, 2).value,
            'Доля': sheet_1.cell(i, 3).value
        })
        d = decimal.Decimal(sheet_1.cell(i, 2).value)
        d = moneyfmt(d, sep=' ')
        table_contents_country.append({
            fo_reg: sheet_1.cell(i, 1).value,
            netto_stoim: d,
            'Доля': sheet_1.cell(i, 3).value
        })
    # сумма долей 4-х крупнейших ФО
    if len(table_contents_country) == 2:
        country_part = float(table_contents_country[0]['Доля'])
    elif len(table_contents_country) == 3:
        country_part = float(table_contents_country[0]['Доля']) + float(table_contents_country[1]['Доля'])
    elif len(table_contents_country) == 4:
        country_part = float(table_contents_country[0]['Доля']) + float(table_contents_country[1]['Доля']) + float(
            table_contents_country[2]['Доля'])
    else:
        country_part = float(table_contents_country[0]['Доля']) + float(table_contents_country[1]['Доля']) + float(
        table_contents_country[2]['Доля'] + float(table_contents_country[3]['Доля']))

    return table_contents_country, country_part, table_contents_country_int


# сводная таблица с двумя столбцами и одной колонкой со значениями (например: федеральный округ, регион и объем импорта)
def pivot_table_fo_reg(path_fo_reg, pivot_index, pivot_values, pivot_1_column, pivot_2_column, path_exel_last):
    data_frame_country = pd.read_excel(path_exel_last)
    report_table = data_frame_country.pivot_table(index=pivot_index, values=pivot_values, aggfunc='sum', margins=True).\
        round(0)
    report_table.to_excel(path_fo_reg, sheet_name='Report')

    table_contents_country = []
    workbook = load_workbook(path_fo_reg)
    sheet_1 = workbook['Report']
    values_fo_last = ''
    for i in range(2, sheet_1.max_row + 1):
        table_year_dict = dict()
        values_fo = sheet_1.cell(i, 1).value
        if values_fo is None:
            table_year_dict[pivot_1_column] = values_fo_last
        else:
            table_year_dict[pivot_1_column] = values_fo
            values_fo_last = values_fo
        table_year_dict[pivot_2_column] = sheet_1.cell(i, 2).value
        for j in range(3, sheet_1.max_column + 1):
            d = sheet_1.cell(i, j).value
            if d is not None:
                d = decimal.Decimal(sheet_1.cell(i, j).value)
                d = moneyfmt(d, sep=' ')
                table_year_dict[sheet_1.cell(1, j).value] = d
        table_contents_country.append(table_year_dict)

    return table_contents_country


# сводная таблица по странам и стоимости
def pivot_country_cost():
    table_contents_country = []
    table_contents_country_int = []
    workbook = load_workbook(r'exel\report_2021.xlsx')
    sheet_1 = workbook['Report']
    for i in range(2, sheet_1.max_row + 1):
        table_contents_country_int.append({
            'Страны': sheet_1.cell(i, 1).value,
            'NETTO': sheet_1.cell(i, 2).value,
            'Доля': sheet_1.cell(i, 3).value
        })
        d = decimal.Decimal(sheet_1.cell(i, 2).value)
        d = moneyfmt(d, sep=' ')
        table_contents_country.append({
            'Страны': sheet_1.cell(i, 1).value,
            'NETTO': d,
            'Доля': sheet_1.cell(i, 3).value
        })
    # сумма долей 4-х крупнейших стран
    country_part = float(table_contents_country[0]['Доля']) + float(table_contents_country[1]['Доля']) + float(
        table_contents_country[2]['Доля'] + float(table_contents_country[3]['Доля']))

    return table_contents_country, country_part, table_contents_country_int


def pivot_country_cost_excel():
    # вычисляем сводную таблицу по стоимости в долларах за последний год по странам
    excel_path = r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"
    excel_file_path = path_exel2

    excel_file = os.path.join(excel_path, excel_file_path)
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False
    excel.Workbooks.Open(excel_file)
    excel.ActiveWorkbook.SaveAs(excel_file, FileFormat=51, ConflictResolution=2)
    excel.DisplayAlerts = True
    excel.ActiveWorkbook.Close()

    data_frame_country = pd.read_excel(path_exel2)

    report_table_country_netto = data_frame_country.pivot_table(index='Страны', values='NETTO', aggfunc='sum').round(0)

    wb2 = openpyxl.Workbook()
    wb2.save(filename=r'exel\report_2021_dol.xlsx')

    with pd.ExcelWriter(r'exel\report_2021_dol.xlsx', mode='a', if_sheet_exists='overlay') as writer:
        report_table_country_netto.to_excel(writer, sheet_name='Report1')

    report_table_country_stoim = data_frame_country.pivot_table(index='Страны', values='STOIM', aggfunc='sum').round(0)

    with pd.ExcelWriter(r'exel\report_2021_dol.xlsx', mode='a', if_sheet_exists='overlay') as writer2:
        report_table_country_stoim.to_excel(writer2, sheet_name='Report2')

    wb5 = load_workbook(r'exel\report_2021_dol.xlsx')
    sheet = wb5['Report1']
    sheet['C1'] = 'Стоимость'
    sheet['D1'] = 'Цена'
    sheet2 = wb5['Report2']
    quantity_row = sheet.max_row
    for k in range(2, quantity_row + 1):
        formula = sheet2['B' + str(k)].value
        formula2 = '=C' + str(k) + '/B' + str(k)
        sheet.cell(row=k, column=3, value=formula)
        sheet.cell(row=k, column=4, value=formula2)
    wb5.save(filename=r'exel\report_2021_dol.xlsx')
    wb5.close()
    pass


def pivot_country_cost_word():
    # вставляем сводную таблицу по стоимости в долларах за последний год по странам
    excel_path = r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"
    excel_file_path = r'D:\PyCharmProject\Marketing_report\exel\report_2021_dol.xlsx'

    excel_file = os.path.join(excel_path, excel_file_path)
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False
    excel.Workbooks.Open(excel_file)
    excel.ActiveWorkbook.SaveAs(excel_file, FileFormat=51, ConflictResolution=2)
    excel.DisplayAlerts = True
    excel.ActiveWorkbook.Close()

    table_contents_country_int = []
    data_frame_country = pd.read_excel(excel_file_path, sheet_name='Report1')
    number_row = len(data_frame_country)
    for i in range(0, number_row):
        if data_frame_country['Цена'][i] != '#ДЕЛ/0!':
            table_contents_country_int.append({
                'Страны': data_frame_country['Страны'][i],
                'Стоимость': round(data_frame_country['Цена'][i], 2)
            })
        else:
            table_contents_country_int.append({
                'Страны': data_frame_country['Страны'][i],
                'Стоимость': 'нет данных'
            })
    return table_contents_country_int


def pivot_table_country_dol():
    pivot1_1('2013')
    pivot_country_cost_excel()

    excel_path = r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"
    excel_file_path = r'D:\PyCharmProject\Marketing_report\exel\report_2021_dol.xlsx'

    excel_file = os.path.join(excel_path, excel_file_path)
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False
    excel.Workbooks.Open(excel_file)
    excel.ActiveWorkbook.SaveAs(excel_file, FileFormat=51, ConflictResolution=2)
    excel.DisplayAlerts = True
    excel.ActiveWorkbook.Close()

    df = pd.read_excel(r'exel\report_2021_dol.xlsx', sheet_name='Report1')
    name_colums1 = df.columns.values.tolist()
    if 'NETTO' in name_colums1:
        df.pop('NETTO')
        df.pop('Стоимость')
        df.loc[:, "Год"] = "2013"
        df.to_excel(r'exel\report_2021_dol2.xlsx', index=False)
    for year in range(1, len(list_year)):
        pivot1_1(list_year[year])
        pivot_country_cost_excel()

        excel_path = r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"
        excel_file_path = r'D:\PyCharmProject\Marketing_report\exel\report_2021_dol.xlsx'

        excel_file = os.path.join(excel_path, excel_file_path)
        excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
        excel.DisplayAlerts = False
        excel.Workbooks.Open(excel_file)
        excel.ActiveWorkbook.SaveAs(excel_file, FileFormat=51, ConflictResolution=2)
        excel.DisplayAlerts = True
        excel.ActiveWorkbook.Close()

        df = pd.read_excel(r'exel\report_2021_dol.xlsx', sheet_name='Report1')
        name_colums = df.columns.values.tolist()
        if 'NETTO' in name_colums:
            df.pop('NETTO')
            df.pop('Стоимость')
            df.loc[:, "Год"] = list_year[year]
            wb_report = load_workbook(r'exel\report_2021_dol2.xlsx')
            sheet = wb_report.active
            row_end = sheet.max_row
            with pd.ExcelWriter(r'exel\report_2021_dol2.xlsx', mode='a', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, startrow=row_end, index=False, header=False)
        else:
            pass

    data_frame_country = pd.read_excel(r'exel\report_2021_dol2.xlsx')
    report_table = data_frame_country.pivot_table(index='Страны', values='Цена', aggfunc='mean', columns='Год',
                                                  margins=True).round(2)
    report_table.to_excel(r'exel\report_2021_dol3.xlsx', sheet_name='Report')

    table_contents_country = []
    workbook = load_workbook(r'exel\report_2021_dol3.xlsx')
    sheet_1 = workbook['Report']
    for i in range(2, sheet_1.max_row + 1):
        table_year_dict = dict()
        table_year_dict['Страны'] = sheet_1.cell(i, 1).value
        for j in range(2, sheet_1.max_column + 1):
            d = sheet_1.cell(i, j).value
            if d is not None:
                table_year_dict[sheet_1.cell(1, j).value] = d
        table_contents_country.append(table_year_dict)
    return table_contents_country


def pivot_country_cost_rub():
    # вычисляем сводную таблицу по стоимости в длолларах за последний год по странам
    excel_path = r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"
    excel_file_path = r'D:\PyCharmProject\Marketing_report\exel\report_2021_dol.xlsx'

    excel_file = os.path.join(excel_path, excel_file_path)
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False
    excel.Workbooks.Open(excel_file)
    excel.ActiveWorkbook.SaveAs(excel_file, FileFormat=51, ConflictResolution=2)
    excel.DisplayAlerts = True
    excel.ActiveWorkbook.Close()

    wb = load_workbook(r'exel\report_2021_dol.xlsx')
    sheet = wb['Report1']
    sheet['E1'] = 'Цена_руб'
    quantity_row = sheet.max_row
    for k in range(2, quantity_row + 1):
        formula2 = '=D' + str(k) + '*' + str(rate[year_now])
        sheet.cell(row=k, column=5, value=formula2)
    wb.save(r'exel\report_2021_dol.xlsx')
    wb.close()

    excel_path = r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"
    excel_file_path = r'D:\PyCharmProject\Marketing_report\exel\report_2021_dol.xlsx'

    excel_file = os.path.join(excel_path, excel_file_path)
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False
    excel.Workbooks.Open(excel_file)
    excel.ActiveWorkbook.SaveAs(excel_file, FileFormat=51, ConflictResolution=2)
    excel.DisplayAlerts = True
    excel.ActiveWorkbook.Close()

    table_contents_country_int = []
    data_frame_country = pd.read_excel(excel_file_path, sheet_name='Report1')
    number_row = len(data_frame_country)
    for i in range(0, number_row):
        if data_frame_country['Цена_руб'][i] != '#ДЕЛ/0!':
            table_contents_country_int.append({
                'Страны': data_frame_country['Страны'][i],
                'Стоимость': round(data_frame_country['Цена_руб'][i], 2)
            })
        else:
            table_contents_country_int.append({
                'Страны': data_frame_country['Страны'][i],
                'Стоимость': 'нет данных'
            })
    return table_contents_country_int
