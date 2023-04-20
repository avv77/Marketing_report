import pymorphy2
from openpyxl import load_workbook
import re

wb = load_workbook(r'D:\Работа_3\production_tems.xlsx')
sheet = wb['all']
quantity_row = sheet.max_row
reg = re.compile('[^а-яА-Я ]')

for i in range(2, quantity_row + 1):
    words_list = sheet['A' + str(i)].value
    words_list_pro = words_list.split()
    words_list_pro2 = []
    for word_123 in words_list_pro:
        word_1 = reg.sub('', word_123)
        words_list_pro2.append(word_1)
    word_list_now = []
    word_gent = 'Рынок'

    for word1 in words_list_pro2:
        morph = pymorphy2.MorphAnalyzer()
        word_morf = morph.parse(word1)[0]
        gent = word_morf.inflect({'gent'})
        try:
            word_pro = gent.word
            word_list_now.append(word_pro)
        except Exception as exp:
            continue
    for word2 in word_list_now:
        word_gent = word_gent + ' ' + word2

    sheet['C' + str(i)].value = word_gent
wb.save(r'D:\Работа_3\production_tems.xlsx')
