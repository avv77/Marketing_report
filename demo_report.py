from docxtpl import DocxTemplate
from openpyxl import load_workbook
import os
import comtypes.client
from PyPDF2 import PdfReader, PdfWriter


def tnved_number2():
    wb = load_workbook(r'exel\1_4_production.xlsx')
    sheet = wb['tnved']
    production_tnved_number = production
    quantity_row = sheet.max_row
    for i in range(2, quantity_row + 1):
        value_production_name = str(sheet['A' + str(i)].value)
        if value_production_name == production_tnved_number:
            value_tnved_name = str(sheet['C' + str(i)].value)
            value_tnved_name_mod = value_tnved_name.split(sep=';')
            len_list_code_name = len(value_tnved_name_mod)
            if len_list_code_name == 1:
                name_tnved1_ = value_tnved_name_mod[0]
                name_tnved2_ = ''
                name_tnved3_ = ''
            elif len_list_code_name == 2:
                name_tnved1_ = value_tnved_name_mod[0]
                name_tnved2_ = value_tnved_name_mod[1]
                name_tnved3_ = ''
            else:
                name_tnved1_ = value_tnved_name_mod[0]
                name_tnved2_ = value_tnved_name_mod[1]
                name_tnved3_ = value_tnved_name_mod[2]
            value_tnved_code = str(sheet['B' + str(i)].value)
            value_tnved_code_mod = value_tnved_code.split(sep=';')
            tnved_1 = value_tnved_code_mod
            len_list_code = len(value_tnved_code_mod)
            if len_list_code == 1:
                tnved1_1 = value_tnved_code_mod[0]
                tnved2_1 = ''
                tnved3_1 = ''
            elif len_list_code == 2:
                tnved1_1 = value_tnved_code_mod[0]
                tnved2_1 = value_tnved_code_mod[1]
                tnved3_1 = ''
            else:
                tnved1_1 = value_tnved_code_mod[0]
                tnved2_1 = value_tnved_code_mod[1]
                tnved3_1 = value_tnved_code_mod[2]
            return name_tnved1_, name_tnved2_, name_tnved3_, tnved_1, tnved1_1, tnved2_1, tnved3_1


number_them = 0
wb = load_workbook(r'exel\1_4_production.xlsx')
sheet = wb['demo']
for i in range(2, 22):
    production_pro = str(sheet['E' + str(i)].value)
    production = str(sheet['A' + str(i)].value)
    name_file = str(sheet['C' + str(i)].value)

    tnved_code_name = tnved_number2()

    tnved1 = tnved_code_name[4]
    tnved2 = tnved_code_name[5]
    tnved3 = tnved_code_name[6]
    name_tnved1 = tnved_code_name[0]
    name_tnved2 = tnved_code_name[1]
    name_tnved3 = tnved_code_name[2]

    context = {'продукция1': production_pro, 'код_вэд1': tnved1, 'код_вэд2': tnved2, 'код_вэд3': tnved3,
               'наим_код_вэд1': name_tnved1, 'наим_код_вэд2': name_tnved2, 'наим_код_вэд3': name_tnved3}

    file_report_pattern = r'D:\PyCharmProject\Marketing_report\doc\report_demo\Импорт_пример_демо.docx'

    doc = DocxTemplate(file_report_pattern)

    doc.render(context)
    file_report_pattern_final = r'D:\PyCharmProject\Marketing_report\doc\report_demo\Импорт_пример_демо_финал.docx'
    doc.save(file_report_pattern_final)
    file_report_pdf1 = os.path.join('demo', f'{name_file}.pdf')

    wdFormatPDF = 17

    in_file = os.path.abspath(file_report_pattern_final)
    out_file = os.path.abspath(file_report_pdf1)
    number_them += 1
    print(f'Обработана тема номер {number_them}')

    word = comtypes.client.CreateObject('Word.Application')
    doc = word.Documents.Open(in_file)
    doc.SaveAs(out_file, FileFormat=wdFormatPDF)
    doc.Close()
    word.Quit()

    reader = PdfReader(out_file)
    writer = PdfWriter()

    writer.append_pages_from_reader(reader)
    metadata = reader.metadata
    writer.add_metadata(metadata)

    # Write your custom metadata here:
    writer.add_metadata({"/Author": "Бюро Готовых Исследований"})

    with open(out_file, 'wb') as fp:
        writer.write(fp)
