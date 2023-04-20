import decimal
import pandas as pd
from openpyxl import load_workbook
from pivot_table_country import moneyfmt
from settings import list_year, year_now


# выгружаем из общей базы по производству данные по продукции
def production_base(production1):
    df = pd.read_excel(r'exel\1_4_production.xlsx', sheet_name='all')
    df1 = df[df["Продукция ОКПД"] == production1]
    df1.to_excel(r'exel\report_production.xlsx', sheet_name='Report', index=False)
    wb = load_workbook(r'exel\report_production.xlsx')
    sheet = wb['Report']
    quantity_row = sheet.max_row
    for i in range(2, quantity_row + 1):
        if sheet['B' + str(i)].value == 'тыс. тонн':
            value_parametr = 'тонн'
            sheet['B' + str(i)].value = value_parametr
            value_2013 = sheet['D' + str(i)].value * 1000
            sheet['D' + str(i)].value = value_2013
            value_2014 = sheet['E' + str(i)].value * 1000
            sheet['E' + str(i)].value = value_2014
            value_2015 = sheet['F' + str(i)].value * 1000
            sheet['F' + str(i)].value = value_2015
            value_2016 = sheet['G' + str(i)].value * 1000
            sheet['G' + str(i)].value = value_2016
            value_2017 = sheet['H' + str(i)].value * 1000
            sheet['H' + str(i)].value = value_2017
            value_2018 = sheet['I' + str(i)].value * 1000
            sheet['I' + str(i)].value = value_2018
            value_2019 = sheet['J' + str(i)].value * 1000
            sheet['J' + str(i)].value = value_2019
            value_2020 = sheet['K' + str(i)].value * 1000
            sheet['K' + str(i)].value = value_2020
            value_2021 = sheet['L' + str(i)].value * 1000
            sheet['L' + str(i)].value = value_2021
        if sheet['B' + str(i)].value == 'Декалитр':
            value_parametr = 'тонн'
            sheet['B' + str(i)].value = value_parametr
            value_2013 = sheet['D' + str(i)].value * 10 / 1000
            sheet['D' + str(i)].value = value_2013
            value_2014 = sheet['E' + str(i)].value * 10 / 1000
            sheet['E' + str(i)].value = value_2014
            value_2015 = sheet['F' + str(i)].value * 10 / 1000
            sheet['F' + str(i)].value = value_2015
            value_2016 = sheet['G' + str(i)].value * 10 / 1000
            sheet['G' + str(i)].value = value_2016
            value_2017 = sheet['H' + str(i)].value * 10 / 1000
            sheet['H' + str(i)].value = value_2017
            value_2018 = sheet['I' + str(i)].value * 10 / 1000
            sheet['I' + str(i)].value = value_2018
            value_2019 = sheet['J' + str(i)].value * 10 / 1000
            sheet['J' + str(i)].value = value_2019
            value_2020 = sheet['K' + str(i)].value * 10 / 1000
            sheet['K' + str(i)].value = value_2020
            value_2021 = sheet['L' + str(i)].value * 10 / 1000
            sheet['L' + str(i)].value = value_2021
        if sheet['B' + str(i)].value == 'Тысяча декалитров':
            value_parametr = 'тонн'
            sheet['B' + str(i)].value = value_parametr
            value_2013 = sheet['D' + str(i)].value * 10
            sheet['D' + str(i)].value = value_2013
            value_2014 = sheet['E' + str(i)].value * 10
            sheet['E' + str(i)].value = value_2014
            value_2015 = sheet['F' + str(i)].value * 10
            sheet['F' + str(i)].value = value_2015
            value_2016 = sheet['G' + str(i)].value * 10
            sheet['G' + str(i)].value = value_2016
            value_2017 = sheet['H' + str(i)].value * 10
            sheet['H' + str(i)].value = value_2017
            value_2018 = sheet['I' + str(i)].value * 10
            sheet['I' + str(i)].value = value_2018
            value_2019 = sheet['J' + str(i)].value * 10
            sheet['J' + str(i)].value = value_2019
            value_2020 = sheet['K' + str(i)].value * 10
            sheet['K' + str(i)].value = value_2020
            value_2021 = sheet['L' + str(i)].value * 10
            sheet['L' + str(i)].value = value_2021
        if sheet['B' + str(i)].value == 'кг':
            value_parametr = 'тонн'
            sheet['B' + str(i)].value = value_parametr
            value_2013 = sheet['D' + str(i)].value / 1000
            sheet['D' + str(i)].value = value_2013
            value_2014 = sheet['E' + str(i)].value / 1000
            sheet['E' + str(i)].value = value_2014
            value_2015 = sheet['F' + str(i)].value / 1000
            sheet['F' + str(i)].value = value_2015
            value_2016 = sheet['G' + str(i)].value / 1000
            sheet['G' + str(i)].value = value_2016
            value_2017 = sheet['H' + str(i)].value / 1000
            sheet['H' + str(i)].value = value_2017
            value_2018 = sheet['I' + str(i)].value / 1000
            sheet['I' + str(i)].value = value_2018
            value_2019 = sheet['J' + str(i)].value / 1000
            sheet['J' + str(i)].value = value_2019
            value_2020 = sheet['K' + str(i)].value / 1000
            sheet['K' + str(i)].value = value_2020
            value_2021 = sheet['L' + str(i)].value / 1000
            sheet['L' + str(i)].value = value_2021
    wb.save(r'exel\report_production.xlsx')
    pass


# удаляем строку с общим объемом ('Российская Федерация')
def production_delete_value(production1):
    df = pd.read_excel(r'exel\report_production.xlsx', sheet_name='Report')
    df2 = df[df.Регионы != 'Российская Федерация']
    new_row_sum = pd.DataFrame({'Продукция ОКПД': [production1], 'Размерность': ['Тонна'],
                                'Регионы': ['Итого'], 2013: [df2[2013].sum()], 2014: [df2[2014].sum()], 2015:
                                    [df2[2015].sum()], 2016: [df2[2016].sum()], 2017: [df2[2017].sum()],
                                2018: [df2[2018].sum()], 2019: [df2[2019].sum()], 2020: [df2[2020].sum()],
                                2021: [df2[2021].sum()]})
    df2 = pd.concat([df2, new_row_sum])
    df2.to_excel(r'exel\report_production.xlsx', sheet_name='Report', index=False)
    pass


# создаем словарь по годам и объемам производства по строке "итого"
def production_dict_value_sum():
    wb = load_workbook(r'exel\report_production.xlsx')
    sheet = wb['Report']

    quantity_row = sheet.max_row
    quantity_column = sheet.max_column

    sum_dict_production = {}

    for column_production in range(4, quantity_column + 1):
        cell_year = sheet.cell(row=quantity_row, column=column_production)
        value_year = round(int(cell_year.value))
        cell_year_name = sheet.cell(row=1, column=column_production)
        value_year_name = cell_year_name.value
        sum_dict_production[value_year_name] = value_year

    parameter = sheet.cell(row=2, column=2).value

    return sum_dict_production, parameter


# создаем словарь по годам и динамикой объемов к предыдущему году
def production_dict_dynamics(dict_production):
    dynamics_dict_production = {}
    for year in range(1, len(list_year)):
        dynamics_year2 = round(int(dict_production[int(list_year[year])]))
        dynamics_year1 = round(int(dict_production[int(list_year[year - 1])]))
        dynamics_year = round(dynamics_year2 / dynamics_year1 * 100 - 100, 1)
        key = int(list_year[year])
        dynamics_dict_production[key] = dynamics_year
    return dynamics_dict_production


# сводная таблица по федеральным округам (Таблица 3.3)
def pivot_pro_1():
    data_frame_fo = pd.read_excel(r'exel\report_production.xlsx')
    data_frame_fo_now = data_frame_fo[data_frame_fo.Регионы != 'Итого']
    data_frame_fo_now.to_excel(r'exel\report_pro_fo.xlsx', sheet_name='Report', index=False)

    data_frame_fo = pd.read_excel(r'exel\report_pro_fo.xlsx')
    values_pro = int(year_now)
    report_table_fo = data_frame_fo.pivot_table(index='Регионы', values=values_pro, aggfunc='sum').round(0)
    report_table_fo.to_excel(r'exel\report_pro_fo_2.xlsx', sheet_name='Report')

    df = pd.read_excel(r'exel\report_pro_fo_2.xlsx')
    bit_depth = 0
    col_netto_list = df[values_pro].tolist()
    exist_count = col_netto_list.count(0)

    while exist_count != 0:
        if bit_depth == 5:
            break
        bit_depth += 1
        report_table = data_frame_fo.pivot_table(index='Регионы', values=values_pro, aggfunc='sum').round(bit_depth)
        report_table.to_excel(r'exel\report_pro_fo_2.xlsx', sheet_name='Report')
        df = pd.read_excel(r'exel\report_pro_fo_2.xlsx')
        col_netto_list = df[values_pro].tolist()
        exist_count = col_netto_list.count(0)

    # сортировка таблицы по весу
    df_netto = pd.read_excel(r'exel\report_pro_fo_2.xlsx')
    final_result = df_netto.sort_values(by=values_pro, ascending=False)
    final_result.to_excel(r'exel\report_pro_fo_2.xlsx', sheet_name='Report', index=False)

    # добавляем "итого" по федеральным округам
    df_country_itog = pd.read_excel(r'exel\report_pro_fo_2.xlsx')
    itog = round(df_country_itog[values_pro].sum())
    wb_report = load_workbook(r'exel\report_pro_fo_2.xlsx')
    sheet = wb_report.active
    quantity_row = sheet.max_row
    sheet['A' + str(quantity_row + 1)] = 'Итого'
    sheet['B' + str(quantity_row + 1)] = itog
    wb_report.save(r'exel\report_pro_fo_2.xlsx')

    # добавляем название столбца "Доля"
    wb_proportion = load_workbook(r'exel\report_pro_fo_2.xlsx')
    sheet_proportion = wb_proportion.active
    sheet_proportion['C1'] = 'Доля'
    sheet_proportion['B1'] = 'Объем'

    # расчитываем долю по каждой стране
    counter_country = 1
    quantity_row = sheet_proportion.max_row
    for i_netto in range(2, quantity_row + 1):
        netto_country = sheet_proportion['B' + str(i_netto)].value
        netto_proportion = round(netto_country / itog * 100, counter_country)

        netto_value = round(sheet_proportion['B' + str(i_netto)].value)

        counter_country = 1
        counter_netto = 0

        while netto_value == 0:
            if counter_netto == 5:
                break
            counter_netto += 1
            netto_value = round(sheet_proportion['B' + str(i_netto)].value, counter_netto)
        sheet_proportion['B' + str(i_netto)].value = netto_value

        while netto_proportion == 0:
            counter_country += 1
            if counter_country == 5:
                break
            netto_proportion = round(netto_country / itog * 100, counter_country)
        sheet_proportion['C' + str(i_netto)].value = netto_proportion

    quantity_row = sheet_proportion.max_row
    sheet_proportion['C' + str(quantity_row)] = round(100, 1)
    wb_proportion.save(r'exel\report_pro_fo_2.xlsx')
    pass


# вставляем сводную таблицу (Таблица 3.3) из экселя в ворд
def pivot_pro_2():
    table_contents_country = []
    workbook = load_workbook(r'exel\report_pro_fo_2.xlsx')
    sheet_1 = workbook['Report']
    for i in range(2, sheet_1.max_row + 1):
        d = decimal.Decimal(sheet_1.cell(i, 2).value)
        d = moneyfmt(d, sep=' ')
        table_contents_country.append({
            'Регионы': sheet_1.cell(i, 1).value,
            'Объем': d,
            'Доля': sheet_1.cell(i, 3).value
        })
    return table_contents_country


# сводная таблица по динамике производства по федеральным округам (Таблица 3.4)
def pivot_table_fo_year():
    data_frame_fo = pd.read_excel(r'exel\report_pro_fo.xlsx')
    data_frame_fo.pop('Продукция ОКПД')
    data_frame_fo.pop('Размерность')
    data_frame_fo.to_excel(r'exel\report_din_fo.xlsx', index=False)
    table_contents_country = []
    workbook = load_workbook(r'exel\report_din_fo.xlsx')
    sheet_1 = workbook.active
    for i in range(2, sheet_1.max_row + 1):
        table_year_dict = dict()
        table_year_dict['ФО'] = sheet_1.cell(i, 1).value
        for j in range(2, sheet_1.max_column + 1):
            d = sheet_1.cell(i, j).value
            if d is not None:
                d = decimal.Decimal(sheet_1.cell(i, j).value)
                d = moneyfmt(d, sep=' ')
                table_year_dict[int(sheet_1.cell(1, j).value)] = d
            else:
                table_year_dict[sheet_1.cell(1, j).value] = ''
        table_contents_country.append(table_year_dict)

    return table_contents_country
