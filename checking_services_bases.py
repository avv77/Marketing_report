from openpyxl import load_workbook
import pandas as pd

path_base = r'D:\Работа\Росстат\Количество сотрудников_2017-2022.xlsx'
path_zar = r'D:\Работа\Росстат\Зарплата_2019-2021.xlsx'
path_ren = r'D:\Работа\Росстат\Рентабельность_2021.xlsx'
path_cost = r'D:\Работа\Росстат\Выручка по отрасли_2017-2022_сентябрь.xlsx'
path_seb = r'D:\Работа\Росстат\Себестоимость_2017-2022.xlsx'
path_prib = r'D:\Работа\Росстат\Прибыль_2017-2022_сентябрь.xlsx'
path_invest = r'D:\Работа\Росстат\Инвестиции в основной капитал_2017-2022.xlsx'


def okved():
    wb = load_workbook(path_base)
    sheet = wb['Данные']
    wb1 = load_workbook(path_zar)
    sheet1 = wb1.active
    wb2 = load_workbook(path_ren)
    sheet2 = wb2.active
    wb3 = load_workbook(path_cost)
    sheet3 = wb3.active
    wb4 = load_workbook(path_seb)
    sheet4 = wb4.active
    wb5 = load_workbook(path_prib)
    sheet5 = wb5.active
    wb6 = load_workbook(path_invest)
    sheet6 = wb6.active
    content_list = []

    for i in range(6316, 7841):
        counter = 0
        value = sheet['B' + str(i)].value
        for row in range(4, 6224):
            value_wb1 = str(sheet1['B' + str(row)].value)
            if value == value_wb1:
                counter += 1
                break
        for row2 in range(4, 56401):
            value_wb2 = str(sheet2['B' + str(row2)].value)
            if value == value_wb2:
                counter += 1
                break
        for row3 in range(4, 32811):
            value_wb3 = str(sheet3['A' + str(row3)].value)
            if value == value_wb3:
                counter += 1
                break
        for row4 in range(4, 32565):
            value_wb4 = str(sheet4['A' + str(row4)].value)
            if value == value_wb4:
                counter += 1
                break
        for row5 in range(4, 33000):
            value_wb5 = str(sheet5['A' + str(row5)].value)
            if value == value_wb5:
                counter += 1
                break
        for row6 in range(6, 53456):
            value_wb6 = str(sheet6['A' + str(row6)].value)
            if value == value_wb6:
                counter += 1
                break
        if counter == 6:
            content_list.append(value)

    df = pd.DataFrame(content_list)
    with pd.ExcelWriter(r'D:\Работа\Темы для сайта\Услуги.xlsx') as writer:
        df.to_excel(writer, 'темы')


okved()
