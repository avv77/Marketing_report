import sqlite3
import pandas as pd
from openpyxl import load_workbook
from settings import path_base, path_exel, production


def production_2017():
    wb = load_workbook(r'D:\Работа3\Производство.xlsx')
    sheet = wb['all']
    sheet2 = wb['2017-2021']
    quantity_row = sheet.max_row
    quantity_row2 = sheet2.max_row
    for i in range(2, quantity_row + 1):
        value1 = sheet['N' + str(i)].value
        if value1 == 1:
            val_okpd2 = sheet['M' + str(i)].value
            reg = sheet['C' + str(i)].value
            for name_prod in range(2, quantity_row2 + 1):
                value_okpd = sheet2['A' + str(name_prod)].value
                reg_okpd = sheet2['C' + str(name_prod)].value
                if value_okpd == val_okpd2 and reg_okpd == reg:
                    value_2017 = sheet2['D' + str(name_prod)].value
                    value_2018 = sheet2['E' + str(name_prod)].value
                    value_2019 = sheet2['F' + str(name_prod)].value
                    value_2020 = sheet2['G' + str(name_prod)].value
                    value_2021 = sheet2['H' + str(name_prod)].value
                    sheet['H' + str(i)].value = value_2017
                    sheet['I' + str(i)].value = value_2018
                    sheet['J' + str(i)].value = value_2019
                    sheet['K' + str(i)].value = value_2020
                    sheet['L' + str(i)].value = value_2021
                    print(f'Обработана запись {name_prod}')
    wb.save(r'D:\Работа3\Производство.xlsx')


def production_south_fo():
    wb = load_workbook(r'D:\Работа3\Производство.xlsx')
    sheet = wb['2017-2021']
    quantity_row = sheet.max_row
    for i in range(2, quantity_row + 1):
        value_fo = sheet['C' + str(i)].value
        value_fo2 = sheet['C' + str(i - 1)].value
        if value_fo == 'Дальневосточный федеральный округ':
            if value_fo2 == 'Дальневосточный федеральный округ':
                value_2018 = sheet['E' + str(i)].value
                value_2019 = sheet['F' + str(i)].value
                value_2020 = sheet['G' + str(i)].value
                value_2021 = sheet['H' + str(i)].value
                sheet['E' + str(i - 1)].value = value_2018
                sheet['F' + str(i - 1)].value = value_2019
                sheet['G' + str(i - 1)].value = value_2020
                sheet['H' + str(i - 1)].value = value_2021
                sheet.delete_rows(i, 1)
                print(f'удалена строка {i}.')
    wb.save(r'D:\Работа3\Производство.xlsx')
    pass


def tnved_check():
    wb = load_workbook(r'D:\Работа3\1_4_production.xlsx')
    sheet = wb['tnved']
    quantity_row = sheet.max_row
    tnved_list = []
    for i in range(2, quantity_row + 1):
        value_tnved = str(sheet['B' + str(i)].value)
        value_tnved_last = str(sheet['B' + str(i - 1)].value)
        if value_tnved != value_tnved_last:
            value_tnved_mod = value_tnved.split(sep=';')
            for number in value_tnved_mod:
                tnved_list.append(number)
    # выгружаем импорт по коду из базы данных таможни
    for number_ved in tnved_list:
        conn = sqlite3.connect(path_base)
        with pd.ExcelWriter(path_exel) as writer:
            cur = conn.cursor()
            cur.execute(f'SELECT * FROM "2021" WHERE NAPR = "ИМ" AND TNVED LIKE "{number_ved}"')
            result = cur.fetchall()
            daf = pd.DataFrame(result)
            daf.to_excel(writer, '2021', index=False)
        wb = load_workbook(path_exel)
        sheet = wb.active
        value_tnved1 = str(sheet['A' + str(2)].value)
        if not value_tnved1:
            print(f'импорт по коду {number_ved} пустой')


def tnved_number(path, product):
    wb = load_workbook(path)
    sheet = wb['tnved']
    production_tnved_number = product
    quantity_row = sheet.max_row
    for i in range(2, quantity_row + 1):
        value_production_name = str(sheet['A' + str(i)].value)
        if value_production_name == production_tnved_number:
            value_tnved_name = str(sheet['C' + str(i)].value)
            value_tnved_name_mod = value_tnved_name.split(sep=';')
            len_list_code_name = len(value_tnved_name_mod)
            if len_list_code_name == 1:
                name_tnved1_ = value_tnved_name_mod[0]
                name_tnved2_ = ''
                name_tnved3_ = ''
            elif len_list_code_name == 2:
                name_tnved1_ = value_tnved_name_mod[0]
                name_tnved2_ = value_tnved_name_mod[1]
                name_tnved3_ = ''
            else:
                name_tnved1_ = value_tnved_name_mod[0]
                name_tnved2_ = value_tnved_name_mod[1]
                name_tnved3_ = value_tnved_name_mod[2]
            value_tnved_code = str(sheet['B' + str(i)].value)
            value_tnved_code_mod = value_tnved_code.split(sep=';')
            tnved_1 = value_tnved_code_mod
            len_list_code = len(value_tnved_code_mod)
            if len_list_code == 1:
                tnved1_1 = value_tnved_code_mod[0]
                tnved2_1 = ''
                tnved3_1 = ''
            elif len_list_code == 2:
                tnved1_1 = value_tnved_code_mod[0]
                tnved2_1 = value_tnved_code_mod[1]
                tnved3_1 = ''
            else:
                tnved1_1 = value_tnved_code_mod[0]
                tnved2_1 = value_tnved_code_mod[1]
                tnved3_1 = value_tnved_code_mod[2]
            return name_tnved1_, name_tnved2_, name_tnved3_, tnved_1, tnved1_1, tnved2_1, tnved3_1
